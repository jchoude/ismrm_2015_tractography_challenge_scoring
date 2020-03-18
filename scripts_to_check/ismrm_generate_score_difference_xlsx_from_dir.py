#! /usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import division, print_function

import argparse
import json
import os
import pickle

from openpyxl import Workbook


DESCRIPTION = """
    Generates an excel spreadsheet showing the differences in scores with
    the classical IB technique vs the QB-based IB technique.
    """


def _create_scores_xlsx(classical_dir, new_dir, subs_fname, out_fname):
    ordered_subs = []
    with open(subs_fname, 'rb') as subs_f:
        for l in subs_f:
            tokens = os.path.splitext(l)[0].strip().split('_')
            team_id = tokens[0]
            sub_id = tokens[1]

            if len(ordered_subs) < int(team_id):
                ordered_subs.append([int(sub_id)])
            else:
                ordered_subs[int(team_id) - 1].append(int(sub_id))

    wb = Workbook()
    cur_sheet = wb.active

    cur_sheet.cell(column=2, row=1).value = "IB"
    cur_sheet.cell(column=3, row=1).value = "IB - new"
    cur_sheet.cell(column=4, row=1).value = "IB ratio"
    cur_sheet.cell(column=6, row=1).value = "IC"
    cur_sheet.cell(column=7, row=1).value = "IC - new"
    cur_sheet.cell(column=8, row=1).value = "IC ratio"
    cur_sheet.cell(column=10, row=1).value = "NC"
    cur_sheet.cell(column=11, row=1).value = "NC - new"
    cur_sheet.cell(column=12, row=1).value = "NC ratio"

    cur_sheet.cell(column=14, row=1).value = "VCWP"
    cur_sheet.cell(column=15, row=1).value = "VCWP - new"

    cur_row = 2

    for team_id, team_content in enumerate(ordered_subs):
        for sub_id in team_content:
            cur_sheet.cell(column=1, row=cur_row).value = \
                "{0}_{1}".format(team_id + 1, sub_id)
            classical_score_fname = os.path.join(classical_dir,
                                                 "{0}_{1}.pkl".format(
                                                         team_id + 1, sub_id))
            new_score_fname = os.path.join(new_dir,
                                           "{0}_{1}.pkl".format(
                                                         team_id + 1, sub_id))

            if not os.path.isfile(classical_score_fname) or not \
                os.path.isfile(new_score_fname):
                cur_sheet.cell(column=2, row=cur_row).value = "Missing score file"
            else:
                with open(classical_score_fname, 'rb') as classical_score_f:
                    classical_scores = pickle.load(classical_score_f)
                with open(new_score_fname, 'rb') as new_score_f:
                    new_scores = pickle.load(new_score_f)

                cur_sheet.cell(column=2, row=cur_row).value = classical_scores['IB']
                cur_sheet.cell(column=3, row=cur_row).value = new_scores['IB']
                cur_sheet.cell(column=4, row=cur_row).value = float(new_scores['IB']) / float(classical_scores['IB'])
                cur_sheet.cell(column=6, row=cur_row).value = classical_scores['IC']
                cur_sheet.cell(column=7, row=cur_row).value = new_scores['IC']
                cur_sheet.cell(column=8, row=cur_row).value = float(new_scores['IC']) / float(classical_scores['IC'])
                cur_sheet.cell(column=10, row=cur_row).value = classical_scores['NC']
                cur_sheet.cell(column=11, row=cur_row).value = new_scores['NC']
                cur_sheet.cell(column=12, row=cur_row).value = float(new_scores['NC']) / float(classical_scores['NC'])

                cur_sheet.cell(column=14, row=cur_row).value = classical_scores['VCWP']
                cur_sheet.cell(column=15, row=cur_row).value = new_scores['VCWP']

            cur_row += 1

    cur_sheet.title = "Scores differences"

    wb.save(out_fname)


def buildArgsParser():

    p = argparse.ArgumentParser(formatter_class=argparse.RawTextHelpFormatter,
                                description=DESCRIPTION)

    p.add_argument('classical_dir', action='store', metavar='CLASSICAL_DIR', type=str,
                   help='Input directory containing the classical scores pickle files.')
    p.add_argument('new_dir', action='store', metavar='NEW_DIR', type=str,
                   help='Input directory containing the new scores pickle files.')
    p.add_argument('subs_list', action='store', metavar='SUBS', type=str,
                   help='file containing the ordered submissions list')
    #p.add_argument('basic_bundles_attribs', action='store', metavar='ATTRIBS',
    #               type=str, help='Path to basic bundles json attribute file')
    p.add_argument('output', action='store', metavar='OUT_FILE',
                   type=str, help='Path to output xlsx file')

    return p


def main():
    parser = buildArgsParser()
    args = parser.parse_args()

    if not os.path.isdir(args.classical_dir):
        parser.error("Classical dir doesn't exist.")

    if not os.path.isdir(args.new_dir):
        parser.error("New dir doesn't exist.")

    if not os.path.isfile(args.subs_list):
        parser.error("Not an input file.")

    # Find all bundle names
    #with open(args.basic_bundles_attribs, 'r') as f:
    #    attribs = json.load(f)

    #bundles_names = [os.path.splitext(s)[0] for s in sorted(attribs.keys())]

    _create_scores_xlsx(args.classical_dir, args.new_dir, args.subs_list,
                        args.output)


if __name__ == "__main__":
    main()
