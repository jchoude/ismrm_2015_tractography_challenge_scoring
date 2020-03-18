#!/usr/bin/env python

from __future__ import division

import argparse
import glob
import json
import os
import logging

import nibabel as nb
import numpy as np

from openpyxl import Workbook

import tractometer.pipeline_helper as helper

###############
# Script part #
###############
DESCRIPTION = 'Script used to compute the overlaps and overreaches between ' \
              'the Ground truth and extracted Valid Bundles.'


def _get_all_overlaps(in_dir, bundles_names):
    subs_list = glob.glob(os.path.join(in_dir, '*.json'))

    overlaps = {}

    for sub_fname in sorted(subs_list):
        sub_name = os.path.splitext(os.path.basename(sub_fname))[0]
        with open(sub_fname, 'rb') as score_f:
            scores = json.load(score_f)

        overlaps[sub_name] = {}
        for bundle_id in bundles_names:
            overlaps[sub_name][bundle_id] = {"overlap": scores['overlap_per_bundle'][bundle_id],
                                             "overreach": scores['overreach_per_bundle'][bundle_id],
                                             "overreach_norm_gt": scores['overreach_norm_gt_per_bundle'][bundle_id]}

    return overlaps


def _write_header(ws, bundles_names):
    for col_enum, bundle_name in enumerate(bundles_names, start=2):
        ws.cell(column=col_enum, row=1).value = bundle_name


def _write_stat(ws, full_stats, stat_key, cur_row, bundles_names):
    for team_id, sub_name in enumerate(sorted(full_stats.keys())):
        ws.cell(column=1, row=cur_row).value = sub_name

        sub_stats = full_stats.get(sub_name)
        if sub_stats is None:
            ws.cell(column=2, row=cur_row).value = "NOT DONE"
        else:
            for col_enum, bundle_name in enumerate(bundles_names, start=2):
                bundle_overlap = sub_stats.get(bundle_name)

                if bundle_overlap is not None:
                    ws.cell(column=col_enum, row=cur_row).value = \
                        bundle_overlap[stat_key]
                else:
                    ws.cell(column=col_enum, row=cur_row).value = 0

        cur_row += 1

    return cur_row


def _save_overlaps(overlaps, out_file, bundles_names):
    wb = Workbook()
    cur_sheet = wb.active
    cur_sheet.title = "Bundles overlap"

    _write_header(cur_sheet, bundles_names)

    cur_row = 2
    cur_row = _write_stat(cur_sheet, overlaps, 'overlap', cur_row, bundles_names)

    # Overreaches sheet
    cur_sheet = wb.create_sheet(title="Overreaches")

    _write_header(cur_sheet, bundles_names)

    cur_row = 2
    cur_row = _write_stat(cur_sheet, overlaps, 'overreach', cur_row, bundles_names)

    # Overreaches normalized by GT bundles volume sheet
    cur_sheet = wb.create_sheet(title="Overreaches over GT")

    _write_header(cur_sheet, bundles_names)

    cur_row = 2
    cur_row = _write_stat(cur_sheet, overlaps, 'overreach_norm_gt',
                          cur_row, bundles_names)

    wb.save(out_file)


def buildArgsParser():
    p = argparse.ArgumentParser(description=DESCRIPTION,
                                formatter_class=argparse.RawTextHelpFormatter)

    p.add_argument('input_dir', action='store', metavar='IN_DIR', type=str,
                   help='Input directory containing scores json files.')
    p.add_argument('basic_bundles_attribs', action='store', metavar='ATTRIBS',
                   type=str, help='Path to basic bundles json attribute file')
    p.add_argument('out_file', action='store', metavar='OUT_FILE',
                   type=str, help='Path to output xlsx file')

    #Other
    p.add_argument('-f', dest='is_forcing', action='store_true',
                   required=False, help='overwrite output files')
    p.add_argument('-v', dest='is_verbose', action='store_true',
                   required=False, help='produce verbose output')

    return p


def main():
    parser = buildArgsParser()
    args = parser.parse_args()

    isForcing = args.is_forcing
    isVerbose = args.is_verbose

    if isVerbose:
        helper.VERBOSE = True
        logging.basicConfig(level=logging.DEBUG)

    if os.path.isfile(args.out_file):
        if isForcing:
            os.remove(args.out_file)
        else:
            parser.error('output file already exists.')

    if not os.path.isdir(args.input_dir):
        parser.error('Not an input dir.')

    # Find all bundle names
    with open(args.basic_bundles_attribs, 'r') as f:
        attribs = json.load(f)

    bundles_names = [os.path.splitext(s)[0] for s in sorted(attribs.keys())]

    overlaps = _get_all_overlaps(args.input_dir, bundles_names)

    _save_overlaps(overlaps, args.out_file, bundles_names)

if __name__ == "__main__":
    main()
