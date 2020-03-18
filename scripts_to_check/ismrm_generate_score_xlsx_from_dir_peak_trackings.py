#! /usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import division, print_function

import argparse
import glob
import json
import os

from openpyxl import Workbook


DESCRIPTION = """
    Generates the full excel spreadsheet containing a the scores of
    all submissions.
    """


def _create_scores_xlsx(in_dir, bundles_names, out_fname):
    ordered_subs = []

    subs_list = glob.glob(os.path.join(in_dir, '*.json'))

    for sub_fname in sorted(subs_list):
        ordered_subs.append(os.path.splitext(os.path.basename(sub_fname))[0])

    wb = Workbook()
    cur_sheet = wb.active

    cur_sheet.cell(column=2, row=1).value = "VB"
    cur_sheet.cell(column=3, row=1).value = "IB"
    cur_sheet.cell(column=4, row=1).value = "VC"
    cur_sheet.cell(column=5, row=1).value = "IC"
    cur_sheet.cell(column=6, row=1).value = "VCWP"
    cur_sheet.cell(column=7, row=1).value = "NC"
    cur_sheet.cell(column=8, row=1).value = "Total streamlines"

    # Write all bundles names
    start_col = 9
    for col_idx, bname in enumerate(bundles_names, start=start_col):
        cur_sheet.cell(column=col_idx, row=1).value = bname

    cur_row = 2
    algo_versions = []

    for sub_id, sub_name in enumerate(ordered_subs):
        cur_sheet.cell(column=1, row=cur_row).value = sub_name
        score_fname = os.path.join(in_dir, "{}.json".format(sub_name))

        if not os.path.isfile(score_fname):
            cur_sheet.cell(column=2, row=cur_row).value = "NOT DONE"
        else:
            with open(score_fname, 'rb') as score_f:
                scores = json.load(score_f)
                algo_versions.append(scores['algo_version'])

            cur_sheet.cell(column=2, row=cur_row).value = scores['VB']
            cur_sheet.cell(column=3, row=cur_row).value = scores['IB']
            cur_sheet.cell(column=4, row=cur_row).value = scores['VC']
            cur_sheet.cell(column=5, row=cur_row).value = scores['IC']
            cur_sheet.cell(column=6, row=cur_row).value = scores['VCWP']
            cur_sheet.cell(column=7, row=cur_row).value = scores['NC']
            cur_sheet.cell(column=8, row=cur_row).value = \
                scores['total_streamlines_count']

            strl_per_bundles = scores['streamlines_per_bundle']

            for col_idx, bname in enumerate(bundles_names, start=start_col):
                cur_sheet.cell(column=col_idx, row=cur_row).value = \
                    strl_per_bundles.get(bname, 0)

        cur_row += 1

    if algo_versions.count(algo_versions[0]) != len(algo_versions):
        raise ValueError('A score file had a different algorithm version than' +
                         'the others.')

    cur_sheet.title = "Scores, technique {}".format(algo_versions[0])

    wb.save(out_fname)


def buildArgsParser():

    p = argparse.ArgumentParser(formatter_class=argparse.RawTextHelpFormatter,
                                description=DESCRIPTION)

    p.add_argument('input_dir', action='store', metavar='IN_DIR', type=str,
                   help='Input directory containing scores json files.')
    p.add_argument('basic_bundles_attribs', action='store', metavar='ATTRIBS',
                   type=str, help='Path to basic bundles json attribute file')
    p.add_argument('output', action='store', metavar='OUT_FILE',
                   type=str, help='Path to output xlsx file')

    return p


def main():
    parser = buildArgsParser()
    args = parser.parse_args()

    if not os.path.isdir(args.input_dir):
        parser.error('Not an input dir.')

    # Find all bundle names
    with open(args.basic_bundles_attribs, 'r') as f:
        attribs = json.load(f)

    bundles_names = [os.path.splitext(s)[0] for s in sorted(attribs.keys())]

    _create_scores_xlsx(args.input_dir, bundles_names, args.output)


if __name__ == "__main__":
    main()
