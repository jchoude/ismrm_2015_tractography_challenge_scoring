#!/usr/bin/env python

from __future__ import division

import argparse
import glob
import json
import os
import logging

import nibabel as nb
import numpy as np

from openpyxl import Workbook

import tractometer.pipeline_helper as helper

###############
# Script part #
###############
DESCRIPTION = 'Script used to compute the overlaps and overreaches between ' \
              'the Ground truth and extracted Valid Bundles.'

def _compute_overlap(basic_data, mask_data):
    basic_non_zero = np.count_nonzero(basic_data)

    overlap = np.logical_and(basic_data, mask_data)
    overlap_count = np.count_nonzero(overlap)

    print('  OL_C: {}, basic_nz: {}, OL: {}'.format(overlap_count, basic_non_zero, overlap_count / basic_non_zero))

    return overlap_count / basic_non_zero


def _compute_overreach(gt_data, candidate_data):
    diff = candidate_data - gt_data
    diff[diff < 0] = 0

    overreach_count = np.count_nonzero(diff)

    return overreach_count / np.count_nonzero(candidate_data)


def _compute_overreach_normalize_gt(gt_data, candidate_data):
    diff = candidate_data - gt_data
    diff[diff < 0] = 0

    overreach_count = np.count_nonzero(diff)

    return overreach_count / np.count_nonzero(gt_data)


def _compute_all_overlaps(masks_dir, basic_masks_dir):
    basic_masks_list = os.listdir(basic_masks_dir)
    bundles_masks = {}
    for basic_mask_f in basic_masks_list:
        bundle_name = os.path.splitext(os.path.splitext(os.path.basename(basic_mask_f))[0])[0]
        bundle_img = nb.load(os.path.join(basic_masks_dir, basic_mask_f))
        bundles_masks[bundle_name] = bundle_img.get_data()

    overlaps = {}

    for mask_file in glob.iglob(os.path.join(masks_dir, '*')):
        chunks = os.path.splitext(os.path.splitext(os.path.basename(mask_file))[0])[0].split('_', 3)
        print(mask_file)

        team_id = int(chunks[0])
        sub_id = int(chunks[1])
        bundle_id = chunks[3]

        full_sub_id = "{0}_{1}".format(team_id, sub_id)

        mask_img = nb.load(mask_file)
        mask_data = mask_img.get_data()

        bundle_overlap = _compute_overlap(bundles_masks[bundle_id], mask_data)
        bundle_overreach = _compute_overreach(bundles_masks[bundle_id], mask_data)
        bundle_overreach_norm_gt = _compute_overreach_normalize_gt(bundles_masks[bundle_id], mask_data)

        bu = overlaps.get(full_sub_id)
        if bu is None:
            overlaps[full_sub_id] = {bundle_id: {"overlap": bundle_overlap,
                                                 "overreach": bundle_overreach,
                                                 "overreach_norm_gt": bundle_overreach_norm_gt}}
        else:
            bu[bundle_id] = {"overlap": bundle_overlap,
                             "overreach": bundle_overreach,
                             "overreach_norm_gt": bundle_overreach_norm_gt}

    return overlaps


def _write_header(ws, bundle_attribs):
    for col_enum, bundle_name in enumerate(sorted(bundle_attribs.keys()), start=2):
        bn = os.path.splitext(bundle_name)[0]
        ws.cell(column=col_enum, row=1).value = bn


def _write_stat(ws, ordered_subs, full_stats, stat_key, cur_row, bundles_attribs):
    for team_id, team_content in enumerate(ordered_subs):
        for sub_id in team_content:
            full_sub_id = "{0}_{1}".format(team_id + 1, sub_id)
            ws.cell(column=1, row=cur_row).value = full_sub_id

            sub_stats = full_stats.get(full_sub_id)
            if sub_stats is None:
                ws.cell(column=2, row=cur_row).value = "NOT DONE"
            else:
                for col_enum, bundle_name in enumerate(sorted(bundles_attribs.keys()), start=2):
                    bn = os.path.splitext(bundle_name)[0]
                    bundle_overlap = sub_stats.get(bn)

                    if bundle_overlap is not None:
                        ws.cell(column=col_enum, row=cur_row).value = \
                            bundle_overlap[stat_key]
                    else:
                        ws.cell(column=col_enum, row=cur_row).value = 0

            cur_row += 1

    return cur_row


def _save_overlaps(overlaps, out_file, subs_fname, bundles_attribs_file):
    ordered_subs = []
    with open(subs_fname, 'rb') as subs_f:
        for l in subs_f:
            tokens = os.path.splitext(l)[0].strip().split('_')
            team_id = tokens[0]
            sub_id = tokens[1]

            if len(ordered_subs) < int(team_id):
                ordered_subs.append([int(sub_id)])
            else:
                ordered_subs[int(team_id) - 1].append(int(sub_id))

    with open(bundles_attribs_file, 'rb') as bf:
        bundles_attribs = json.load(bf)

    wb = Workbook()
    cur_sheet = wb.active
    cur_sheet.title = "Bundles overlap"

    _write_header(cur_sheet, bundles_attribs)

    cur_row = 2
    cur_row = _write_stat(cur_sheet, ordered_subs, overlaps, 'overlap', cur_row,
                          bundles_attribs)

    # Overreaches sheet
    cur_sheet = wb.create_sheet(title="Overreaches")

    _write_header(cur_sheet, bundles_attribs)

    cur_row = 2
    cur_row = _write_stat(cur_sheet, ordered_subs, overlaps, 'overreach', cur_row,
                          bundles_attribs)

    # Overreaches normalized by GT bundles volume sheet
    cur_sheet = wb.create_sheet(title="Overreaches over GT")

    _write_header(cur_sheet, bundles_attribs)

    cur_row = 2
    cur_row = _write_stat(cur_sheet, ordered_subs, overlaps, 'overreach_norm_gt',
                          cur_row, bundles_attribs)

    wb.save(out_file)


def buildArgsParser():
    p = argparse.ArgumentParser(description=DESCRIPTION,
                                formatter_class=argparse.RawTextHelpFormatter)

    p.add_argument('masks_dir', action='store',
                   metavar='DIR', type=str, help='Directory with all VB masks')

    p.add_argument('basic_masks_dir',   action='store',
                   metavar='DIR',  type=str,
                   help='directory containing the masks of the basic bundles')

    p.add_argument('subs_list', action='store', metavar='SUBS', type=str,
                   help='file containing the ordered submissions list')

    p.add_argument('basic_bundles_attribs', action='store', metavar='ATTRIBS',
                   type=str, help='Path to basic bundles json attribute file')

    p.add_argument('out_file', action='store', metavar='out_excel',
                   type=str, help='output excel file containing the bundle overlap')

    #Other
    p.add_argument('-f', dest='is_forcing', action='store_true',
                   required=False, help='overwrite output files')
    p.add_argument('-v', dest='is_verbose', action='store_true',
                   required=False, help='produce verbose output')

    return p


def main():
    parser = buildArgsParser()
    args = parser.parse_args()

    isForcing = args.is_forcing
    isVerbose = args.is_verbose

    if isVerbose:
        helper.VERBOSE = True
        logging.basicConfig(level=logging.DEBUG)

    if not os.path.isdir(args.masks_dir):
        parser.error('"{0}" must be a dir!'.format(args.masks_dir))

    if not os.path.isdir(args.basic_masks_dir):
        parser.error('"{0}" must be a directory!'.format(args.basic_masks_dir))

    if os.path.isfile(args.out_file):
        if isForcing:
            os.remove(args.out_file)
        else:
            parser.error('output file already exists.')

    overlaps = _compute_all_overlaps(args.masks_dir, args.basic_masks_dir)

    _save_overlaps(overlaps, args.out_file, args.subs_list,
                   args.basic_bundles_attribs)

if __name__ == "__main__":
    main()
